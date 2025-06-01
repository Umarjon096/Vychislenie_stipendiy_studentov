import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font


def generator(work_folder, save_folder, output_file):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    ultra_red_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')


    fayly = [f for f in os.listdir(work_folder) if f.lower().endswith('.xlsx') and f != output_file]
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for fayl in fayly:
            print(work_folder + '/' + fayl)
            wb = load_workbook(work_folder + '/' + fayl, data_only=True)
            ws = wb.active
            group_code = str(ws['B4'].value).strip().split('-')[0]
            print(f"Обрабатывается: {group_code} из {fayl}")

            meta = pd.read_excel(work_folder + '/' + fayl, header=None, nrows=5)
            df_raw = pd.read_excel(work_folder + '/' + fayl, header=None)
            start_idx = df_raw[df_raw.iloc[:, 0] == '№'].index[0]
            df = pd.read_excel(work_folder + '/' + fayl, header=start_idx)

            start_col = 3
            end_col = start_col
            for col in df.columns[start_col:]:
                if pd.isna(col) or str(col).strip() == '':
                    break
                end_col += 1

            insert_pos = end_col
            for new_col in ['Qarz', 'Baho', 'Stipendiya']:
                df.insert(insert_pos, new_col, '')
                insert_pos += 1

            grade_cols = df.columns[start_col:insert_pos - 3]

            for idx, row in df.iterrows():
                grades_cleaned = []
                for col in grade_cols:
                    val = str(row[col]) if pd.notna(row[col]) else ''
                    num_str = re.sub(r'\s*\[\d+\]', '', val).strip()
                    try:
                        num = float(num_str)
                        grades_cleaned.append(num)
                    except:
                        grades_cleaned.append(None)

                qarz = sum(1 for g in grades_cleaned if g is None or g < 60)
                df.at[idx, 'Qarz'] = qarz

                if qarz == 0 and grades_cleaned:
                    min_score = min(grades_cleaned)
                    if 90 <= min_score <= 100:
                        baho = 5
                    elif 70 <= min_score <= 89:
                        baho = 4
                    elif 60 <= min_score <= 69:
                        baho = 3
                    else:
                        baho = ''
                    df.at[idx, 'Baho'] = baho
                else:
                    df.at[idx, 'Baho'] = ''

            meta.to_excel(writer, sheet_name=group_code, index=False, header=False, startrow=0)
            df.to_excel(writer, sheet_name=group_code, index=False, startrow=len(meta) + 1)

    wb_result = load_workbook(output_file)

    summary_data = []
    qarz_counter = 0
    baho_counter = {3: 0, 4: 0, 5: 0}
    student_counter = 0

    for sheet in wb_result.worksheets:
        if sheet.title == "Hisobot":
            continue

        for cell in ['A1', 'A2', 'A3', 'A4']:
            sheet[cell].value = None

        sheet.page_margins = PageMargins(left=0.16, right=0.16, top=0.4, bottom=0.4)
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

        for col_idx in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_idx == 1:
                sheet.column_dimensions[col_letter].width = 3
            elif col_idx == 2:
                sheet.column_dimensions[col_letter].width = 42
            elif col_idx == 3:
                sheet.column_dimensions[col_letter].width = 17
            elif col_idx == sheet.max_column - 2:
                sheet.column_dimensions[col_letter].width = 5.5
            elif col_idx == sheet.max_column - 1:
                sheet.column_dimensions[col_letter].width = 5.5
            elif col_idx == sheet.max_column:
                sheet.column_dimensions[col_letter].width = 11
            else:
                col_name = sheet[f"{col_letter}7"].value
                if col_name:
                    sheet.column_dimensions[col_letter].width = 5 + len(str(col_name)) / 12
                else:
                    sheet.column_dimensions[col_letter].width = 12

        for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                cell.border = thin_border
                if isinstance(cell.value, str) and 'Davlat granti' in cell.value:
                    cell.fill = blue_fill
                # if isinstance(cell.value, str):
                if cell.value == '' or cell.value == None:
                    if cell.col_idx < sheet.max_column -1:
                        cell.fill = ultra_red_fill
                if isinstance(cell.value, str):
                    if re.search(r'\[1\]', cell.value):
                        try:
                            vval = int(cell.value.split('[')[0].replace(' ', ''))
                            if 90 <= vval <= 100:
                                cell.fill = red_fill
                                cell.font = Font(color="9C0006")
                            elif 70 <= vval <= 89:
                                cell.fill = green_fill
                                cell.font = Font(color="006100")
                        except:
                            print("pass")
                        continue
                    elif '[' in cell.value and ']' in cell.value:
                        cell.fill = yellow_fill

        qarz_col_idx = sheet.max_column - 2
        baho_col_idx = sheet.max_column - 1
        stipend_col_idx = sheet.max_column

        qarz_col_letter = get_column_letter(qarz_col_idx)
        baho_col_letter = get_column_letter(baho_col_idx)
        stipend_col_letter = get_column_letter(stipend_col_idx)

        group_name = sheet["B4"].value.strip()

        for row_idx in range(8, sheet.max_row + 1):
            fio = sheet[f"B{row_idx}"].value
            grand = sheet[f"C{row_idx}"].value
            qarz_val = sheet[f"{qarz_col_letter}{row_idx}"].value
            baho_val = sheet[f"{baho_col_letter}{row_idx}"].value
            stipend = None

            student_counter += 1
            if qarz_val and qarz_val != 0:
                qarz_counter += 1
            if baho_val in [3, 4, 5]:
                baho_counter[baho_val] += 1

            payment_type = sheet[f"C{row_idx}"].value
            if qarz_val and qarz_val != 0:
                sheet[f"{stipend_col_letter}{row_idx}"].value = ''
                summary_data.append([len(summary_data) + 1, fio, grand, group_name, qarz_val, baho_val, ''])
                continue

            if payment_type == 'Davlat granti':
                if baho_val == 5: stipend = 5
                if baho_val == 4: stipend = 4
                if baho_val == 3:
                    grades_cleaned = []
                    for col_idx in range(4, sheet.max_column - 3 + 1):
                        val = sheet[f"{get_column_letter(col_idx)}{row_idx}"].value
                        if val is None:
                            continue
                        num_str = re.sub(r'\s*\[\d+\]', '', str(val)).strip()
                        try:
                            num = float(num_str)
                            grades_cleaned.append(num)
                        except:
                            continue
                    count_3 = sum(1 for g in grades_cleaned if 60 <= g <= 69)
                    percent_3 = (count_3 / len(grades_cleaned)) * 100 if grades_cleaned else 0
                    stipend = 3 if percent_3 < 30 else None

            else:
                if baho_val == 5:
                    stipend = '20%'

            cell = sheet[f"{stipend_col_letter}{row_idx}"]
            cell.alignment = Alignment(horizontal="center")
            if stipend == '20%':
                cell.value = 0.2
                cell.number_format = '0%'
            else:
                cell.value = stipend

            summary_data.append([
                len(summary_data) + 1, fio, grand, group_name, qarz_val, baho_val, cell.value
            ])

    # Создаём новый лист Umumiy
    summary_ws = wb_result.create_sheet("Hisobot")
    headers = ['№', 'F. I. O.', "To‘lov shakli", 'Guruhi', 'Qarz', 'Baho', 'Stipendiya']
    for col_idx, header in enumerate(headers, start=1):
        cell = summary_ws.cell(row=1, column=col_idx, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.font = Font(bold=True)

    for row_idx, row_data in enumerate(summary_data, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = summary_ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx != 2:
                cell.alignment = Alignment(horizontal='center')
            if col_idx == 7 and cell.value == 0.2:
                cell.number_format = '0%'

    # После записи данных — вычисляем и вставляем статистику по qarz и baho
    qarz_col = 5  # 'qarz'
    baho_col = 6  # 'baho'

    # Объединяем ячейки H1:I1 и K1:L1
    summary_ws.merge_cells("I1:J1")
    summary_ws.merge_cells("L1:M1")

    # Записываем заголовки
    summary_ws["I1"] = "Qarzdorlar"
    summary_ws["L1"] = "Baholanganlar"

    # Выравнивание и граница
    summary_ws["I1"].alignment = Alignment(horizontal="center")
    summary_ws["L1"].alignment = Alignment(horizontal="center")

    summary_ws["I1"].font = Font(bold=True)
    summary_ws["L1"].font = Font(bold=True)

    # I2:I13 — "1 ta qarz", ..., "12 ta qarz" + подсчёт количества
    qarz_counts = {i: 0 for i in range(1, 13)}
    total_qarzdor = 0
    for row in range(2, summary_ws.max_row + 1):
        val = summary_ws.cell(row=row, column=qarz_col).value
        if isinstance(val, int) and 1 <= val <= 12:
            qarz_counts[val] += 1
            total_qarzdor += 1

    for i in range(1, 13):
        summary_ws[f"I{i + 1}"] = f"{i} ta qarz"
        summary_ws[f"J{i + 1}"] = qarz_counts[i]
        summary_ws[f"J{i + 1}"].alignment = Alignment(horizontal="center")

    summary_ws["I14"] = "Umumiy soni"
    summary_ws["J14"] = total_qarzdor
    summary_ws["J14"].alignment = Alignment(horizontal="center")

    # Baholanganlar
    baho_counts = {3: 0, 4: 0, 5: 0}
    total_baholangan = 0
    for row in range(2, summary_ws.max_row + 1):
        val = summary_ws.cell(row=row, column=baho_col).value
        if val in [3, 4, 5]:
            baho_counts[val] += 1
            total_baholangan += 1

    summary_ws["L2"] = "3 lar"
    summary_ws["L3"] = "4 lar"
    summary_ws["L4"] = "5 lar"
    summary_ws["L5"] = "Umumiy soni"

    summary_ws["M2"] = baho_counts[3]
    summary_ws["M3"] = baho_counts[4]
    summary_ws["M4"] = baho_counts[5]
    summary_ws["M5"] = total_baholangan

    for x in range(2, 6):
        summary_ws[f"M{x}"].alignment = Alignment(horizontal="center")

    # summary_ws["M2"].alignment = Alignment(horizontal="center")
    # summary_ws["M3"].alignment = Alignment(horizontal="center")
    # summary_ws["M4"].alignment = Alignment(horizontal="center")
    # summary_ws["M5"].alignment = Alignment(horizontal="center")

    summary_ws.column_dimensions['A'].width = 5
    summary_ws.column_dimensions['B'].width = 46
    summary_ws.column_dimensions['C'].width = 17
    summary_ws.column_dimensions['D'].width = 20
    summary_ws.column_dimensions['E'].width = 10
    summary_ws.column_dimensions['F'].width = 10
    summary_ws.column_dimensions['G'].width = 13
    summary_ws.column_dimensions['I'].width = 13
    summary_ws.column_dimensions['J'].width = 8
    summary_ws.column_dimensions['L'].width = 13
    summary_ws.column_dimensions['M'].width = 13

    # Граница для блока H1:I14 (Qarzdorlar)
    for row in range(1, 15):  # H1 to H14
        for col in range(9, 11):  # H=8, I=9
            cell = summary_ws.cell(row=row, column=col)
            cell.border = thin_border

    # Граница для блока J1:K5 (Baholanganlar)
    for row in range(1, 6):  # J1 to J5
        for col in range(12, 14):  # J=10, K=11
            cell = summary_ws.cell(row=row, column=col)
            cell.border = thin_border

    # output = save_folder + '/' + output_file
    # output_filename = "itog.xlsx"
    output_path = os.path.join(save_folder, output_file)
    wb_result.save(output_path)
    print("✅ Готово: файл обновлен и лист 'Umumiy' добавлен.")
